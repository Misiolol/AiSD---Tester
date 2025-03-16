import subprocess
import os
import time
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

def read_setup_config(setup_file):
    config = {}
    if os.path.isfile(setup_file):
        with open(setup_file, "r") as f:
            for line in f:
                line = line.strip()
                if '=' in line:
                    key, value = line.split('=')
                    config[key.strip()] = value.strip()
    else:
        print(f"Error: '{setup_file}' not found.")
    return config

def run_executable_with_multiple_inputs(executable, input_files, output_excel_file, output_chart_files, plots_enabled, iteration):
    if not os.path.isfile(executable):
        print(f"Error: '{executable}' not found.")
        return

    results = {"Random": [], "V-Type": [], "A-Type": []}

    for index, input_file in enumerate(input_files):
        if not os.path.isfile(input_file):
            print(f"Error: Input file '{input_file}' not found. Skipping...\n")
            continue

        print(f"Running '{executable}' with input from '{input_file}'...\n")

        with open(input_file, "r") as infile:
            input_lines = infile.readlines()

        input_data = "".join(input_lines)
        run_cmd = [f"./{executable}"] if os.name != "nt" else [executable]

        start_time = time.time()
        run_process = subprocess.run(run_cmd, input=input_data, capture_output=True, text=True)
        end_time = time.time()
        execution_time = round(end_time - start_time, 6)  
        num_lines = len(input_lines)

        
        print("Program Output:\n", run_process.stdout)
        print("Error Output (if any):\n", run_process.stderr)
        print(f"Execution Time: {execution_time:.6f} seconds")
        print("-" * 50)

        category = "Random" if "R" in input_file else "V-Type" if "V" in input_file else "A-Type"
        results[category].append({
            "File Name": input_file,
            "Number of Numbers in File": num_lines,
            f"Iteration {iteration}": execution_time  
        })

    # Save or update results in the respective Excel files for each category
    for category, data in results.items():
        output_file = f"./user-outputs/{category.lower()}_execution_results.xlsx"
        
        if os.path.exists(output_file):
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                book = load_workbook(output_file)

                df_new = pd.DataFrame(data)
                
                if category in book.sheetnames:
                    existing_df = pd.read_excel(output_file, sheet_name=category)

                    for _, row in df_new.iterrows():
                        match_row = existing_df["File Name"] == row["File Name"]
                        if match_row.any():
                            idx = existing_df.index[match_row][0]  
                            existing_df.loc[idx, f"Iteration {iteration}"] = row[f"Iteration {iteration}"]
                        else:
                            existing_df = pd.concat([existing_df, pd.DataFrame([row])], ignore_index=True)
                    
                    existing_df.to_excel(writer, sheet_name=category, index=False)
                else:
                    df_new.to_excel(writer, sheet_name=category, index=False)
        else:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=category, index=False)

        print(f"Results for iteration {iteration} saved to '{output_file}'.")

    if plots_enabled:
        for category, data in results.items():
            if data:
                df = pd.DataFrame(data)
                plt.figure(figsize=(10, 6))
                plt.plot(df["Number of Numbers in File"], df[f"Iteration {iteration}"], marker='o', linestyle='-')
                plt.title(f"Execution Time vs. Number of Numbers ({category}) - Iteration {iteration}")
                plt.xlabel("Number of Numbers in File")
                plt.ylabel("Execution Time (seconds)")
                plt.grid(True)
                plt.savefig(output_chart_files[category].replace(".png", f"_iter{iteration}.png"))
                print(f"{category} plot for iteration {iteration} saved to '{output_chart_files[category]}'.")
                
def main():
    setup_file = "setup.py"
    config = read_setup_config(setup_file)

    random_enabled = config.get('random', 'False') == 'True'
    Vtype_enabled = config.get('Vtype', 'False') == 'True'
    Atype_enabled = config.get('Atype', 'False') == 'True'
    plots_enabled = config.get('plots', 'False') == 'True'

    input_files = []
    if random_enabled:
        input_files.extend([f"./DataSets/RandomNumbers/numbers_{n}.txt" for n in [10, 100, 400, 800, 1000, 4000, 8000, 10000, 40000, 80000, 100000]])
    if Vtype_enabled:
        input_files.extend([f"./DataSets/V_typesets/v_shaped_{n}.txt" for n in [10, 100, 400, 800, 1000, 4000, 8000, 10000, 40000, 80000, 100000]])
    if Atype_enabled:
        input_files.extend([f"./DataSets/A_typesets/a_shaped_{n}.txt" for n in [10, 100, 400, 800, 1000, 4000, 8000, 10000, 40000, 80000, 100000]])

    output_chart_files = {
        "Random": "./user-outputs/random_execution_time_chart.png",
        "V-Type": "./user-outputs/vtype_execution_time_chart.png",
        "A-Type": "./user-outputs/atype_execution_time_chart.png"
    }

    executable_file = "a.exe"
    reroll_script = "./DataSets/reroll.py"

    for iteration in range(1, 11):
        print(f"\n===== Running Iteration {iteration} =====\n")
        run_executable_with_multiple_inputs(executable_file, input_files, None, output_chart_files, plots_enabled, iteration)
        
        if os.path.isfile(reroll_script):
            print(f"\nRunning '{reroll_script}' to regenerate input data...\n")
            subprocess.run(["python", reroll_script])
        else:
            print(f"Error: '{reroll_script}' not found. Skipping reroll.\n")

if __name__ == "__main__":
    main()
